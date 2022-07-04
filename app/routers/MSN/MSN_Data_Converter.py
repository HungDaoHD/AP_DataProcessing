import pandas as pd
import openpyxl
import numpy as np
import re
import io


class QMeFileConvert:

    def __init__(self):
        pass
        # self.dfData = pd.DataFrame()
        # self.dictData, self.dictVarLbl, self.dictValLbl = dict(), dict(), dict()


    def convert(self, file):
        xlsx = io.BytesIO(file.file.read())
        wb = openpyxl.load_workbook(xlsx)

        wsData = wb['Data']
        wsQres = wb['Question']

        mergedCells = list()
        for group in wsData.merged_cells.ranges:
            mergedCells.append(group)

        for group in mergedCells:
            wsData.unmerge_cells(str(group))

        wsData.delete_rows(1, 4)
        wsData.delete_rows(2, 1)

        for icol in range(1, wsData.max_column + 1):
            if wsData.cell(row=1, column=icol).value is None:
                wsData.cell(row=1,
                            column=icol).value = f'{wsData.cell(row=1, column=icol - 1).value}_{wsData.cell(row=2, column=icol).value}'

        wsData.delete_rows(2, 1)

        data = wsData.values
        columns = next(data)[0:]
        dfData = pd.DataFrame(data, columns=columns)

        lstDrop = [
            'Approve',
            'Reject',
            'Re - do request',  'Re-do request',
            'Reason to reject',
            'Memo'
            'No.',
            'Date',
            'ID',
            'Country',
            'Channel',
            'Chain / Type',
            'Distributor',
            'Method',
            'Panel ID',
            'Panel FB',
            'Panel Email',
            'Panel Phone',
            'Panel Age',
            'Panel Gender',
            'Panel Area',
            'Panel Income',
            'Login ID',
            'User name',
            'Store ID',
            'Store Code',
            'Store name',
            'Store level',
            'District',
            'Ward',
            'Store address',
            'Area group',
            'Store ranking',
            'Region 2',
            'Nhóm cửa hàng',
            'Nhà phân phối',
            'Manager',
            'Telephone number',
            'Contact person',
            'Email',
            'Others 1',
            'Others 2',
            'Others 3',
            'Others 4',
            'Check in',
            'Store Latitude',
            'Store Longitude',
            'User Latitude',
            'User Longitude',
            'Check out',
            'Distance',
            'Task duration',

            'InterviewerID',
            'InterviewerName',
            'RespondentName',
            'RespondentEmail',
            'RespondentPhone',
            'RespondentCellPhone',
            'RespondentAddress',

            'Memo',
            'No.'
        ]

        for col in dfData.columns:
            if col in lstDrop or '_Images' in col:
                dfData.drop(col, inplace=True, axis=1)

        data = wsQres.values
        columns = next(data)[0:]
        dfQres = pd.DataFrame(data, columns=columns)

        wb.close()

        dictQres = dict()
        for idx in dfQres.index:

            strMatrix = (
                '' if dfQres.loc[idx, 'Question(Matrix)'] is None else f"{dfQres.loc[idx, 'Question(Matrix)']}_")
            strNormal = (dfQres.loc[
                             idx, 'Question(Normal)'] if strMatrix == '' else f"{strMatrix}{dfQres.loc[idx, 'Question(Normal)']}")
            strQreName = str(dfQres.loc[idx, 'Name of items'])
            strQreName = strQreName.replace('Rank_', 'Rank') if 'Rank_' in strQreName else strQreName

            dictQres[strQreName] = {
                'type': dfQres.loc[idx, 'Question type'],
                'label': f'{strNormal}',
                'isMAMatrix': True if strMatrix != '' and dfQres.loc[idx, 'Question type'] == 'MA' else False,
                'cats': {
                }
            }

            lstHeaderCol = list(dfQres.columns)
            lstHeaderCol.remove('Name of items')
            lstHeaderCol.remove('Question type')
            lstHeaderCol.remove('Question(Matrix)')
            lstHeaderCol.remove('Question(Normal)')

            for col in lstHeaderCol:
                # if col not in ['Name of items', 'Question type', 'Question(Matrix)', 'Question(Normal)'] \
                #         and dfQres.loc[idx, col] is not None and len(dfQres.loc[idx, col]) > 0:
                if dfQres.loc[idx, col] is not None and len(str(dfQres.loc[idx, col])) > 0:
                    dictQres[strQreName]['cats'].update({str(col): self.cleanhtml(str(dfQres.loc[idx, col]))})

        lstMatrixHeader = list()
        for k in dictQres.keys():
            if dictQres[k]['isMAMatrix'] and len(dictQres[k]['cats'].keys()):
                lstMatrixHeader.append(k)

        if len(lstMatrixHeader):
            for i in lstMatrixHeader:
                for code in dictQres[i]['cats'].keys():
                    lstLblMatrixMA = dictQres[f'{i}_{code}']['label'].split('_')
                    dictQres[f'{i}_{code}']['cats'].update({'1': self.cleanhtml(lstLblMatrixMA[1])})
                    dictQres[f'{i}_{code}']['label'] = f"{dictQres[i]['label']}_{lstLblMatrixMA[1]}"

        dictVarLbl = dict()
        dictValLbl = dict()
        for col in dfData.columns:
            if col in dictQres.keys():
                dictVarLbl[col] = self.cleanhtml(dictQres[col]['label'])

                dictValLbl[col] = dictQres[col]['cats']

            else:
                dictVarLbl[col] = col

        dfData.replace({None: np.nan}, inplace=True)
        dictData = dfData.to_dict('records')

        return dictData, dictVarLbl, dictValLbl




    @staticmethod
    def cleanhtml(raw_html):
        CLEANR = re.compile('<.*?>|&([a-z0-9]+|#[0-9]{1,6}|#x[0-9a-f]{1,6});')
        cleantext = re.sub(CLEANR, '', raw_html)
        return cleantext