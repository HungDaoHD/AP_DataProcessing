import pandas as pd
import openpyxl
import pyreadstat
import io
import numpy as np
import zipfile
import re


class APDataConverter:

    def __init__(self):

        self.strFileName = None
        self.xlsx = None

        self.dfData = pd.DataFrame
        self.dictVarLbl, self.dictValLbl = dict(), dict()

        self.dfQres = pd.DataFrame
        self.dictQres = dict()

        self.savName = None
        self.savFile = None

        self.strMRSet = str()
        self.spsName = None
        self.spsFile = None

        self.zipName = None
        self.zipFile = None


    def load(self, file):

        self.strFileName = file.filename
        self.xlsx = io.BytesIO(file.file.read())

        wb = openpyxl.load_workbook(self.xlsx)

        wsData = wb['Data']
        wsQres = wb['Question']

        mergedCells = list()
        for group in wsData.merged_cells.ranges:
            mergedCells.append(group)

        for group in mergedCells:
            wsData.unmerge_cells(str(group))


        wsData.delete_rows(1, 4)
        wsData.delete_rows(2, 1)

        for icol in range(1, wsData.max_column + 1):
            if wsData.cell(row=1, column=icol).value is None:
                wsData.cell(row=1,
                            column=icol).value = f'{wsData.cell(row=1, column=icol - 1).value}_{wsData.cell(row=2, column=icol).value}'

        wsData.delete_rows(2, 1)


        data = wsData.values
        columns = next(data)[0:]
        dfData = pd.DataFrame(data, columns=columns)

        lstDrop = [
            'Approve',
            'Reject',
            'Re - do request', 'Re-do request',
            'Reason to reject',
            'Memo',
            'No.',
            'Date',
            'Country',
            'Channel',
            'Chain / Type',
            'Distributor',
            'Method',
            'Panel FB',
            'Panel Email',
            'Panel Phone',
            'Panel Age',
            'Panel Gender',
            'Panel Area',
            'Panel Income',
            'Login ID',
            'User name',
            'Store ID',
            'Store Code',
            'Store name',
            'Store level',
            'District',
            'Ward',
            'Store address',
            'Area group',
            'Store ranking',
            'Region 2',
            'Nhóm cửa hàng',
            'Nhà phân phối',
            'Manager',
            'Telephone number',
            'Contact person',
            'Email',
            'Others 1',
            'Others 2',
            'Others 3',
            'Others 4',
            'Check in',
            'Store Latitude',
            'Store Longitude',
            'User Latitude',
            'User Longitude',
            'Check out',
            'Distance',
            'Task duration',

            'Panel ID',
            'InterviewerID',
            'InterviewerName',
            'RespondentName',

            'Edited',
            'Edited by',
            'Edited ratio',
        ]

        for col in dfData.columns:
            if col in lstDrop or '_Images' in col:
                dfData.drop(col, inplace=True, axis=1)


        data = wsQres.values
        columns = next(data)[0:]
        dfQres = pd.DataFrame(data, columns=columns)
        dfQres.replace({np.nan: None}, inplace=True)

        wb.close()


        dictQres = dict()
        for idx in dfQres.index:

            strMatrix = (
                '' if dfQres.loc[idx, 'Question(Matrix)'] is None else f"{dfQres.loc[idx, 'Question(Matrix)']}")
            strNormal = dfQres.loc[idx, 'Question(Normal)']
            strQreName = str(dfQres.loc[idx, 'Name of items'])
            strQreName = strQreName.replace('Rank_', 'Rank') if 'Rank_' in strQreName else strQreName

            dictQres[strQreName] = {
                'type': dfQres.loc[idx, 'Question type'],
                'label': f'{strMatrix}_{strNormal}' if strMatrix != '' else f'{strNormal}',
                'isMAMatrix': True if strMatrix != '' and dfQres.loc[idx, 'Question type'] == 'MA' else False,
                'cats': {}
            }

            for col in dfQres.columns:
                if col not in ['Name of items', 'Question type', 'Question(Matrix)', 'Question(Normal)'] \
                        and dfQres.loc[idx, col] is not None:
                    dictQres[strQreName]['cats'].update({int(col): self.cleanhtml(str(dfQres.loc[idx, col]))})


        lstMatrixHeader = list()
        for k in dictQres.keys():
            if dictQres[k]['isMAMatrix'] and len(dictQres[k]['cats'].keys()):
                lstMatrixHeader.append(k)


        for i in lstMatrixHeader:
            for code in dictQres[i]['cats'].keys():
                strQreLbl = dictQres[i]['label']
                strAttLbl = str(dictQres[f'{i}_{code}']['label'])
                strAttLbl = strAttLbl.replace(strQreLbl.rsplit('_', 1)[0], '')[1:]
                strAttLbl = self.cleanhtml(strAttLbl)

                dictQres[f'{i}_{code}']['label'] = f"{strQreLbl}_{strAttLbl}"

                dictQres[f'{i}_{code}']['cats'].update({1: strAttLbl})

        variable_labels = dict()
        variable_value_labels = dict()


        for col in dfData.columns:

            if col in dictQres.keys():

                variable_labels[col] = self.cleanhtml(dictQres[col]['label'])
                variable_value_labels[col] = dictQres[col]['cats']

            else:
                variable_labels[col] = col


        dfData.replace({None: np.nan}, inplace=True)


        self.dfData = dfData
        self.dictVarLbl, self.dictValLbl = variable_labels, variable_value_labels

        self.dfQres = dfQres
        self.dictQres = dictQres



    def toSav(self):

        savName = self.strFileName.replace('xlsx', 'sav')
        savFile = pyreadstat.write_sav(self.dfData, savName,
                                       column_labels=list(self.dictVarLbl.values()),
                                       variable_value_labels=self.dictValLbl)

        self.savName, self.savFile = savName, savFile



    def getMRSetSyntax(self):

        dictMACols = dict()

        for key, val in self.dictQres.items():
            if val['type'] == 'MA' and key in self.dfData.columns:

                lstColName = str(key).rsplit('_', 1)

                if lstColName[0] in dictMACols.keys():
                    dictMACols[lstColName[0]]['vars'].append(key)
                else:
                    dictMACols[lstColName[0]] = {
                        'name': lstColName[0],
                        'lbl': f"{lstColName[0]}. {val['label'].rsplit('_', 1)[0]}",
                        'vars': [key],
                    }

        temp = """
        *{}.
        MRSETS
          /MDGROUP NAME=${} 
          LABEL='{}'
          CATEGORYLABELS=COUNTEDVALUES
          VARIABLES={}
          VALUE=1
          /DISPLAY NAME=[${}].
        """

        strMRSet = '.'
        for key, val in dictMACols.items():
            strMRSet += temp.format(val['name'], val['name'], val['lbl'], ' '.join(val['vars']), val['name'])


        self.strMRSet = strMRSet

        self.spsName = self.strFileName.replace('xlsx', 'sps')

        with open(f'{self.spsName}', 'w') as text_file:
            text_file.write(strMRSet)

        self.spsFile = text_file




    def zipfiles(self):

        self.zipName = f"{self.strFileName.replace('xlsx', 'zip')}"

        with zipfile.ZipFile(self.zipName, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
            zf.write(self.savName)
            zf.write(self.spsName)

        self.zipFile = zf


    @staticmethod
    def cleanhtml(raw_html):
        CLEANR = re.compile('<.*?>|&([a-z0-9]+|#[0-9]{1,6}|#x[0-9a-f]{1,6});')
        cleantext = re.sub(CLEANR, '', raw_html)
        return cleantext








