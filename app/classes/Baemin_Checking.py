from builtins import Exception

import pandas as pd
import openpyxl
import io
import numpy as np
import openpyxl.styles
from datetime import datetime
import re

TEMP_VAL = 999999999


class BaeminCheck:

    def __init__(self):

        self.strFileName = None
        self.xlsx = None


    def load(self, file):

        self.strFileName = str(file.filename).replace('.xlsx', '_Output.xlsx')
        self.xlsx = io.BytesIO(file.file.read())


    def check(self):

        wb = openpyxl.load_workbook(self.xlsx, data_only=True)

        wsData = wb['Data']
        wsCheck = wb.copy_worksheet(wsData)

        wsCheck.delete_rows(1, 1)

        data = wsCheck.values
        columns = next(data)[0:]
        dfCheck = pd.DataFrame(data, columns=columns)

        wb.remove(wsCheck)

        dfCheck = dfCheck.dropna(how='all')

        dfCheck.replace({np.nan: None}, inplace=True)
        dfCheck.replace({None: 'NULL'}, inplace=True)

        dfCheck['App'] = [np.nan if str(a).upper() not in ['GRAB', 'SHOPEEFOOD', 'GOJEK'] else a for a in dfCheck['App']]

        dfCheck['Date'] = [np.nan if not self.validateDateFormat(a) else a for a in dfCheck['Date']]

        validTime = re.compile(r'^(([0-1]\d)|(2[0-3]))h[0-5]\d$')
        dfCheck['Time'] = [np.nan if not validTime.match(a) else a for a in dfCheck['Time']]

        dfCheck['Area'] = [np.nan if not self.validateDistrict(a) else a for a in dfCheck['Area']]

        dfCheck['AFV range'] = [np.nan if not self.validate_AFV_range(a, b) else a for a, b in zip(dfCheck['AFV range'], dfCheck['AFV'])]

        dfCheck['AOV'] = [np.nan if not self.validateAOV(a, [b, c, d]) else a for a, b, c, d in
                          zip(dfCheck['AOV'], dfCheck['AFV'], dfCheck['Delivery fee'], dfCheck['Applicable fee'])]


        # AFV
        dfCheck = self.validateAFV(dfCheck)

        # Food discount
        dfCheck = self.validate_Food_Discount(dfCheck)

        dfCheck['% Discount/AFV'] = [np.nan if not self.validate_pct_Discount_AFV(a, b, c) else a for a, b, c in
                                     zip(dfCheck['% Discount/AFV'], dfCheck['Food discount'], dfCheck['AFV'])]

        dfCheck['Gi???m gi?? m??n ??n (coupon)'] = [np.nan if not self.validateCoupon(a) else a for a in dfCheck['Gi???m gi?? m??n ??n (coupon)']]

        dfCheck['Delivery fee'] = [np.nan if not self.validate_Fee(a) else a for a in dfCheck['Delivery fee']]

        dfCheck['Applicable fee'] = [np.nan if not self.validate_Fee(a) else a for a in dfCheck['Applicable fee']]

        dfCheck['Delivery discount'] = [np.nan if not self.validate_discount(a, b, c) else a for a, b, c in
                                        zip(dfCheck['Delivery discount'], dfCheck['Delivery fee'], dfCheck['Applicable fee'])]

        dfCheck['Shopee Xu Discount'] = [np.nan if not self.validate_XuDisc(a) else a for a in dfCheck['Shopee Xu Discount']]

        # T???ng gi???m ????n h??ng
        dfCheck = self.validate_Total_Discount(dfCheck)

        # AOV after discount
        dfCheck = self.validate_AOV_After_Discount(dfCheck)

        dfCheck['Merchant'] = [np.nan if not self.validate_Merchant(a) else a for a in dfCheck['Merchant']]

        # No. items
        dfCheck = self.validate_No_Items(dfCheck)

        # Quantity & AFV
        dfCheck = self.validate_Quantity_AFV(dfCheck)


        dfCheck = dfCheck[dfCheck.isna().any(axis=1)][:]

        my_red = openpyxl.styles.colors.Color(rgb='FF0000')
        my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)

        for idx in dfCheck.index:
            for i, v in enumerate(list(dfCheck.columns)):

                if v is None:
                    continue

                if pd.isnull(dfCheck.at[idx, v]):
                    irow = idx + 3
                    icol = i + 1

                    wsData.cell(irow, 3).fill = my_fill
                    wsData.cell(1, icol).fill = my_fill
                    wsData.cell(irow, icol).fill = my_fill





        wb.save(self.strFileName)
        wb.close()


    @staticmethod
    def validateDateFormat(strDate):
        strFormat = '%d/%m/%Y'

        try:
            return bool(datetime.strptime(strDate, strFormat))
        except ValueError:
            return False
        except Exception:
            return False


    @staticmethod
    def validateDistrict(strDist):
        try:

            lstDist = [
                'Qu???n 1',
                'Qu???n 2',
                'Qu???n 3',
                'Qu???n 4',
                'Qu???n 5',
                'Qu???n 6',
                'Qu???n 7',
                'Qu???n 8',
                'Qu???n 9',
                'Qu???n 10',
                'Qu???n 11',
                'Qu???n 12',
                'Qu???n T??n B??nh',
                'Qu???n Ph?? Nhu???n',
                'Qu???n G?? V???p',
                'Qu???n B??nh Th???nh',
                'Qu???n B??nh T??n',
                'Qu???n T??n Ph??',
                'Th??nh ph??? Th??? ?????c',
                'Huy???n H??c M??n',
                'Huy???n C??? Chi',
                'Huy???n Nh?? B??',
                'Huy???n B??nh Ch??nh',
                'Huy???n C???n Gi???',
            ]

            if str(strDist).strip() in lstDist:
                return True
            else:
                return False

        except Exception:
            return False


    @staticmethod
    def validate_AFV_range(strAFVRng, intAFV):
        try:
            if not isinstance(intAFV, (float, int)):
                return False

            if intAFV > 160000:
                strCheck = '160+'
            elif intAFV > 120000:
                strCheck = '120-160'
            elif intAFV > 80000:
                strCheck = '80-120'
            elif intAFV > 60000:
                strCheck = '60-80'
            elif intAFV > 40000:
                strCheck = '40-60'
            elif intAFV <= 40000:
                strCheck = 'Under 40'
            else:
                strCheck = ''

            if strAFVRng != strCheck:
                return False
            else:
                return True

        except Exception:
            return False


    @staticmethod
    def validateAOV(a, lst):

        lst = list(map(lambda x: float(str(x).replace('NULL', '0')), lst))

        try:
            if a != sum(lst):
                return False
            else:
                return True

        except TypeError:
            return False
        except Exception:
            return False


    @staticmethod
    def validateAFV(dfCheck):

        df = dfCheck.copy()

        try:
            df.replace({'NULL': 0}, inplace=True)

            df['AFV'] = [np.nan if not float(a).is_integer() else a for a in df['AFV']]

            df['AFV_Sum'] = [0] * df.shape[0]

            for idx in df.index:

                for i in range(1, 21):

                    if isinstance(df.loc[idx, f'AFV {i}'], (int, float)):
                        df.loc[idx, 'AFV_Sum'] += df.loc[idx, f'AFV {i}']
                    else:
                        df.loc[idx, f'AFV {i}'] = np.nan

            df['AFV'] = [np.nan if a != b else a for a, b in zip(df['AFV'], df['AFV_Sum'])]

            df.loc[df['AFV'] < 1000, ['AFV']] = np.nan

            df['AFV'] = [np.nan if 0 > b - a > 60000 else a for a, b in zip(df['AFV'], df['AOV'])]

            dfCheck['AFV'] = df['AFV']

        except Exception:
            dfCheck['AFV'] = [np.nan] * dfCheck.shape[0]


        return dfCheck


    @staticmethod
    def validate_Food_Discount(dfCheck):

        df = dfCheck.copy()

        try:

            df.loc[df['Food discount'] == 'NULL', ['Food discount']] = TEMP_VAL

            df['Food discount'] = [np.nan if not float(a).is_integer() else a for a in df['Food discount']]

            df.loc[df['Food discount'] < 1000, ['Food discount']] = np.nan

            df['Food discount'] = [np.nan if TEMP_VAL != a > b else a for a, b in zip(df['Food discount'], df['AFV'])]

            df.loc[df['Food discount'] == TEMP_VAL, ['Food discount']] = 'NULL'

            dfCheck['Food discount'] = df['Food discount']

        except Exception:
            dfCheck['Food discount'] = [np.nan] * dfCheck.shape[0]

        return dfCheck


    @staticmethod
    def validate_pct_Discount_AFV(a, b, c):
        try:
            if b == 'NULL':
                b = 0

            if a == b/c:
                return True
            else:
                return False

        except Exception:
            return False


    @staticmethod
    def validateCoupon(val):
        try:
            if val != 'NULL':

                if not isinstance(val, (int, float)):
                    return False

                if val < 500:
                    return False

                if not float(val).is_integer():
                    return False

            return True

        except Exception:
            return False

    @staticmethod
    def validate_Fee(val):
        try:
            if val != 'NULL':

                if not isinstance(val, (int, float)):
                    return False

                if val < 1000:
                    return False

                if val > 60000:
                    return False

                if not float(val).is_integer():
                    return False

            return True

        except Exception:
            return False


    @staticmethod
    def validate_discount(a, b, c):
        try:
            if a != 'NULL':

                if not isinstance(a, (int, float)):
                    return False

                if a < 500:
                    return False

                if not float(a).is_integer():
                    return False

                if pd.isnull(a) or pd.isnull(b) or pd.isnull(c):
                    return False

                if a > float(str(b).replace('NULL', '0')) + float(str(c).replace('NULL', '0')):
                    return False

            return True

        except Exception:
            return False


    @staticmethod
    def validate_XuDisc(val):
        try:
            if val != 'NULL':

                if not isinstance(val, (int, float)):
                    return False

                if val < 100:
                    return False

                if not float(val).is_integer():
                    return False

            return True

        except Exception:
            return False


    @staticmethod
    def validate_Total_Discount(dfCheck):

        df = dfCheck.copy()

        try:
            df.loc[df['T???ng gi???m ????n h??ng'] == 'NULL', ['T???ng gi???m ????n h??ng']] = TEMP_VAL

            df.loc[df['Food discount'] == 'NULL', ['Food discount']] = 0
            df.loc[df['Gi???m gi?? m??n ??n (coupon)'] == 'NULL', ['Gi???m gi?? m??n ??n (coupon)']] = 0
            df.loc[df['Delivery discount'] == 'NULL', ['Delivery discount']] = 0
            df.loc[df['Shopee Xu Discount'] == 'NULL', ['Shopee Xu Discount']] = 0

            df['Total_Discount_Check'] = df['Food discount'] + df['Gi???m gi?? m??n ??n (coupon)'] + df['Delivery discount'] + df['Shopee Xu Discount']

            df.loc[(df['Total_Discount_Check'] != df['T???ng gi???m ????n h??ng']) & (df['T???ng gi???m ????n h??ng'] != TEMP_VAL), ['T???ng gi???m ????n h??ng']] = np.nan

            df.loc[df['T???ng gi???m ????n h??ng'] == TEMP_VAL, ['T???ng gi???m ????n h??ng']] = 'NULL'

            dfCheck['T???ng gi???m ????n h??ng'] = df['T???ng gi???m ????n h??ng']

        except Exception:
            dfCheck['T???ng gi???m ????n h??ng'] = [np.nan] * dfCheck.shape[0]

        return dfCheck


    @staticmethod
    def validate_AOV_After_Discount(dfCheck):

        df = dfCheck.copy()

        try:

            df.loc[df['AOV after discount'] == 'NULL', ['AOV after discount']] = TEMP_VAL

            df.loc[df['AOV'] == 'NULL', ['AOV']] = 0
            df.loc[df['Food discount'] == 'NULL', ['Food discount']] = 0
            df.loc[df['Gi???m gi?? m??n ??n (coupon)'] == 'NULL', ['Gi???m gi?? m??n ??n (coupon)']] = 0
            df.loc[df['Delivery discount'] == 'NULL', ['Delivery discount']] = 0
            df.loc[df['Shopee Xu Discount'] == 'NULL', ['Shopee Xu Discount']] = 0

            df['AOV_after_discount'] = df['AOV'] - df['Food discount'] - df['Gi???m gi?? m??n ??n (coupon)'] - df['Delivery discount'] - df['Shopee Xu Discount']

            df.loc[(df['AOV after discount'] != df['AOV_after_discount']) & (df['AOV after discount'] != TEMP_VAL), ['AOV after discount']] = np.nan
            df.loc[df['AOV after discount'] < 0, ['AOV after discount']] = np.nan

            df.loc[df['AOV after discount'] == TEMP_VAL, ['AOV after discount']] = 'NULL'

            dfCheck['AOV after discount'] = df['AOV after discount']

        except Exception:
            dfCheck['AOV after discount'] = [np.nan] * dfCheck.shape[0]

        return dfCheck


    @staticmethod
    def validate_Merchant(val):

        try:

            lst = [
                '-18???C',
                '3 R??u - G?? R??n, Pizza & Tr?? s???a',
                'A L??? B??n ?????u - B??n ch??? H?? N???i',
                'A Ng??o - Tr?? s???a & ??n v???t',
                'A Ph??n - C??m chi??n & H??? ti???u x??o',
                'A T??i - B??n th???t n?????ng',
                'A T??ng - B??nh m?? b?? n?????ng b?? Cambodia',
                'Ahacook',
                '???m th???c Phan Rang',
                '??n v???t S??i G??n 1992',
                'Ande - C??m da g?? m???t ong',
                'Anh Ba G???o',
                'AZ Tea',
                'B?? B???c - B??nh tr??ng cu???n tr???n',
                'B?? B???y - B??n b??',
                'Balzar De Caf??',
                'B??nh b???ch tu???c Takoyaki Sami',
                'B??nh bao Th??? Ph??t',
                'B??nh canh cua L???c V???ng',
                'B??nh cu???n Ba Mi???n',
                'B??nh cu???n gia truy???n H?? N???i',
                'B??nh Deli',
                'B??nh m?? 1 ph??t 30 gi??y',
                'B??nh m?? A V??',
                'B??nh m?? An An',
                'B??nh m?? BM',
                'B??nh m?? Chim Ch???y',
                'B??nh m?? H?? N???i',
                'B??nh m?? Hu???nh Hoa',
                'B??nh m?? M?? R??m',
                'B??nh M?? ??i',
                'B??nh m?? PewPew',
                'B??nh m?? que Ph??p',
                'B??nh m?? Th??? Nh?? K??? An Na',
                'B??nh m?? Tu???n M???p',
                'B??nh tr??ng ch???m C?? G??nh',
                'B??nh tr??ng Deli & Xi??n que',
                'B??nh Tr??ng N?????ng ???? L???t - ???????ng S??? 8',
                'B??nh tr??ng tr???n th???t b???m Quang Huy',
                'Bathucha - B??nh chu???i Th??i',
                'Beno - M??? ?? s???t b?? M???',
                'Bento Delichi',
                'B??? gi?? qu??n',
                'B?? Kho & C?? Ri G?? H????ng Nga',
                'Bonchon Chicken',
                'B???t - Healthy & Weight Loss Food',
                'Bready - B??nh m?? t????i Burger ????a bay',
                'B??n b?? 229',
                'B??n b?? 31A',
                'B??n b?? ??nh Th????ng',
                'B??n b?? C???u Hai',
                'B??n b?? ????ng Ba Gia H???i',
                'B??n b?? Hoa L??m',
                'B??n b?? hu??? 14B',
                'B??n b?? hu??? C?? Ba',
                'B??n b?? hu??? ????ng Ba',
                'B??n b?? Thi??n L??',
                'B??n cay Th??i 2 Thu???n',
                'B??n ch??? H?? N???i 1982',
                'B??n ch??? H??? G????m',
                'B??n ch??? s???a Nha Trang ',
                'B??n ch??? s???a Nha Trang M???n',
                'B??n ?????u Homemade',
                'B??n ?????u m???m t??m A Ch???nh',
                'B??n ?????u m???m t??m M???t',
                'B??n Ri??u B???p Tr?????ng - B??nh Kh???t V??ng T??u',
                'B??n ri??u b??n b?? C?? Lan',
                'B??n ri??u canh b??n 30',
                'B??n ri??u cua C?? Tuy???n',
                'B??n ri??u cua ???c 66',
                'B??n ri??u g??nh B???n Th??nh',
                'B??n ri??u Nguy???n C???nh Ch??n',
                'B??n ri??u O Nhi',
                'B??n th??i h???i s???n V??n Tr?????ng',
                'B??n th??i ngon Mu???i ???t Xanh',
                'B??n th??i v?? b??n m???m Dung',
                'B??n th???t n?????ng C?? T??n',
                'B??n th???t n?????ng H???ng ??n',
                'B??n th???t n?????ng Ki???u B???o',
                'B??n x????ng b?? O Thanh',
                'Burger King',
                'Busan Korean Food',
                'C?? ri g?? 1357',
                'Caf?? Amazon',
                'Caf?? IYO - c?? ph?? kem mu???i',
                'Canopee',
                'Caztus Ice Blended',
                'Chang Hi',
                'Ch??o dinh d?????ng Vi???t Soup',
                'Ch??o ???ch Singapore Vi???t Sing',
                'Ch??o l??ng - B??n m???c - B??n v???t Ph????ng Nghi',
                'Ch??o s?????n B?? Hi???n',
                'Ch??o s?????n C?? Giang',
                'Ch??o Thu???n Vi???t - Ch??o Dinh D?????ng',
                'Ch?? B?????i ?????ng Th??p',
                'Ch?? b?????i V??nh Long',
                'Ch?? kh??c b???ch Thanh',
                'Ch?? Ngon 3N',
                'Ch?? t??u h??? B?? B???ng TS',
                'Cheese Coffee',
                'Chu Dimsum House',
                'Chuk Chuk - Tr?? V?? C?? ph??',
                'ChungChun Korean Hotdog',
                'Chuti Korean Food',
                'C?? Hai Qu??n - B??n C?? Nha Trang & Nem N?????ng Nha Trang',
                'C??? M??y- C??m V??n Ph??ng & B??n Ch??? C??',
                'C?? Ph????ng - C??m Chi??n & Nui x??o',
                'Coco summer - Tr??i c??y nhanh',
                'Cocoboba - N?????c d???a tr??n ch??u d???a',
                'Coconino - Tea & Cheese',
                'C??m 79',
                'C??m chay Di???u Vy',
                'C??m chi??n Linh ????ng',
                'C??m ch?? F??ng',
                'C??m g?? - Ch??o ???ch Singapore 68',
                'C??m g?? ?????i N??o',
                'C??m g?? ????? Nh???t',
                'C??m g?? H???i Nam',
                'C??m g?? Tam K???',
                'C??m g?? T??n H???i Nam',
                'C??m g?? x???i m??? 142',
                'C??m g?? x???i m??? N??ng N???u',
                'C??m g?? x???i m??? qu??n C?? Ba',
                'C??m g?? x???i m??? Th???ch Lam',
                'C??m nh?? Ph??? Th???',
                'C??m Ni??u Ph????ng B???c',
                'C??m ni??u Thi??n L??',
                'C??m S??i G??n',
                'C??m t???m b???i S??i G??n',
                'C??m t???m Cali',
                'C??m t???m Ch??? Hai',
                'C??m t???m C?? Hoa',
                'C??m t???m Kim Ti???n 2',
                'C??m t???m Kim Ti???n 3',
                'C??m t???m L??ng',
                'C??m t???m Long Xuy??n',
                'C??m t???m Mai',
                'C??m t???m M??y',
                'C??m t???m Minh Long',
                'C??m t???m Ng?? Quy???n',
                'C??m t???m Nh???',
                'C??m t???m Ni M???p',
                'C??m t???m Ph??c L???c Th???',
                'C??m t???m Thu???n Ki???u',
                'Con G?? M??i - C??m g?? Ph?? Y??n',
                'Con S??i - S???a t????i tr??n ch??u ???????ng ??en',
                'C???ng C?? Ph??',
                '?????i K?? Qu??n - C??m G?? Ph?? Y??n & M?? X??o',
                'Daily - H??? ti???u & c?? ph??',
                '?????u M Mix - ?????u N??nh & Rau M??',
                '??en ???? Caf??',
                'D?? Ba - Nem n?????ng Nha Trang',
                'D?? B???y - B??n M???m N??m ???? N???ng',
                'Domino\'s Pizza',
                'Double Tea',
                '?????c K?? m?? gia',
                'Effoc Coffee',
                'FOX Tea - Tr?? s???a & ??n v???t',
                'Funny Beef',
                'G?? C?? B???p',
                'G?? Delichi - g?? l??n m??m & g?? b?? x??i',
                'G?? n?????ng C??i Bang',
                'G?? N?????ng ?? ?? O',
                'G?? r??n Chicken Plus',
                'G?? ta Ngon S??? 1 - C??m g?? & Ch??o g?? & G???i',
                'GAM Coffee - Arabica L???c L??? Lem',
                'GENKI',
                'G??? Cafe',
                'G??c B??nh M?? Ch???o',
                'Gong Cha',
                'Guchi - Burger - Tokbokki Kimbap V?? C??m Tr???n',
                'Gusto Food & Drink',
                'Guta Caf??',
                'Hai Tea - Tea & Coffee',
                'Hamburger B?? mi???ng Hapi',
                'Hancook Korea fast food',
                'Hanuri - Qu??n ??n H??n Qu???c',
                'Haocha Milk Tea',
                'Helios - Ti???m ??n v???t 1999',
                'Highlands Coffee',
                'Highlands Coffee',
                'H???ng tr?? Ng?? Gia',
                'Hot & Cold - Tr?? s???a & Xi??n que',
                'HP C??m t???m',
                'H????ng K?? 9 - C??m g??',
                'Jollibee',
                'Kem b?? - Tr??i c??y t?? 251',
                'Kem B?? 251',
                'KFC',
                'Khoai Bistro - B??nh M?? Ch???o, Beefsteak & M?? ??',
                'Kimbap City',
                'KOI Th??',
                'Laha Cafe & Tr?? S???a',
                'Lavida Coffee & Tea',
                'Loca - B??nh Canh C?? L??c',
                'Lotteria',
                'McDonald\'s',
                'M?? cay Seoul',
                'M?? ???c h???n D?? Lan',
                'M?? tr???n T??n L???a',
                'M?? X?? M???ng Ch?? Cu???i ',
                'M?? ?? Double B',
                'Milano Coffee',
                'M???c V??? Qu??n - C??m d??a n??ng & m?? qu???ng',
                'Neca Fresh Yogurt',
                'Ohzee - B??nh b??o & b??nh b???t l???c',
                'Otoke chicken',
                'Panda Coffee & Tea Express',
                'Papa Chicken',
                'PapaX???t - C??m X??o & B?? B??t T???t',
                'Passio Coffee',
                'Ph??t K?? - C??m Chi??n, M?? X??o & H??? Ti???u X??o',
                'Ph??? 24',
                'Ph??? b?? C?? Trang',
                'Ph??? b?? ??an Ph?????ng ',
                'Ph??c Long',
                'Pizza 4P\'s',
                'Pizza Hut',
                'Popeyes',
                'Qu??n ??n ?????c ??n 86',
                'Qu??n c??m Ng???c H??n',
                'Qu??n c??m Nguy???n K??',
                'Qu??n Con G?? M??i - C??m g?? Ph?? Y??n',
                'Qu??n H??? Ti???u M?? Ho??nh 008',
                'Qu??n H??? Ti???u M???c C?? ??t Saigon',
                'R&B Tea',
                'R??u Cam - G?? r??n, Pizza & Tr?? s???a',
                'Rau M?? Mix',
                'Rau M?? Pha S??i G??n 1982',
                'Royaltea',
                'Sasin - M?? cay 7 c???p ????? H??n Qu???c',
                'Shilin',
                'Shilin - C??m G?? - G?? r??n, Tr?? ????o ????i Loan',
                'S??ng S??nh Milk Tea',
                'Starbucks Coffee',
                'S???a chua tr??n ch??u H??? Long',
                'Subin Steak',
                'S???i C???o 193',
                'Sukiya',
                'S??p cua M???c',
                'S??p cua Thu Hi???n',
                'Susu\'s Burger & B??nh Tr??ng Long An',
                'T??n Ph?????c Takoyaki',
                'T??u H?? Xe Lam',
                'Texas Chicken',
                'Th??nh ?????t - H??? ti???u Nam Vang',
                'Thanh H????ng - B??nh X??o, B??nh Kh???t, B??n M???m Mi???n T??y',
                'The "MY QU???NG" House & Coffee Bar',
                'The 1988 - Tr?? s???a & ??n v???t',
                'The Alley - Tr?? s???a ????i Loan',
                'The Coffee Cean & Tea Leaf',
                'The Coffee House',
                'Th??? Gi???i C??m T???m - B??n Th???t N?????ng - B??nh M?? Th???t N?????ng',
                'The Hideout Tea & Coffee',
                'The Pizza Company',
                'Th??m - Tr?? S???a Ch???ng Kh???u Nghi???p',
                'Ti???n H???i Qu??n - B??n ?????u m???m t??m',
                'Tiger Sugar - ???????ng N??u S???a ????i Loan',
                'Toco Toco Bubble Tea',
                'Tous Le Jours',
                'Tr?? d??u Mr.?????',
                'Tr?? s???a - H???ng tr?? Ng?? Gia',
                'Tr?? S???a & Milo Tutimi',
                'Tr?? s???a A Ng??o',
                'Tr?? s???a Ai Cha',
                'Tr?? s???a Bobapop',
                'Tr?? s???a Comebuy',
                'Tr?? s???a Dhi',
                'Tr?? S???a Fozu',
                'Tr?? s???a Fulong Tea',
                'Tr?? s???a Heekcaa ',
                'Tr?? s???a Huy\'s',
                'Tr?? s???a MayCha',
                'Tr?? s???a Miutea',
                'Tr?? s???a M???c',
                'Tr?? s???a N???ng',
                'Tr?? s???a Sacha',
                'Tr?? s???a Sunday Basic',
                'Tr?? s???a Te Amo',
                'Tr?? s???a Ti??n H?????ng',
                'Tr?? s???a Tr??n 94',
                'Tr??m M?? Tr???n',
                'Vitamin Bar - Sinh T???',
                'Wow Chicken G?? R??n & C??m G??',
                'X??i B??nh Ti??n',
                'X???p - M??n H???n & ???c ',
                'Young Coffee & Tea',
            ]

            lstUpper = [str(i).upper() for i in lst]

            if str(val).upper() not in lstUpper:
                return False

            return True

        except Exception:
            return False


    @staticmethod
    def validate_No_Items(dfCheck):

        df = dfCheck.copy()

        # try:
        df['No_items'] = [0] * df.shape[0]

        for idx in df.index:
            for i in range(1, 21):

                if isinstance(df.loc[idx, f'Quantity {i}'], (int, float)) or df.loc[idx, f'Quantity {i}'] == 'NULL':

                    if df.loc[idx, f'Quantity {i}'] == 'NULL':
                        df.loc[idx, f'Quantity {i}'] = 0

                    df.loc[idx, 'No_items'] = df.loc[idx, 'No_items'] + df.loc[idx, f'Quantity {i}']

                else:
                    df.loc[idx, f'Quantity {i}'] = np.nan



            if not isinstance(df.loc[idx, 'No. items'], (int, float)):
                df.loc[idx, 'No. items'] = np.nan


        df.loc[(df['No. items'] == 0) | (df['No. items'] > 50), ['No. items']] = np.nan
        df.loc[(df['No. items'] != df['No_items']), ['No. items']] = np.nan


        dfCheck['No. items'] = df['No. items']

        # except Exception:
        #
        #     dfCheck['No. items'] = [np.nan] * dfCheck.shape[0]


        return dfCheck


    @staticmethod
    def validate_Quantity_AFV(dfCheck):

        df = dfCheck.copy()

        for i in range(1, 21):
            try:
                df.loc[df[f'Quantity {i}'] == 'NULL', [f'Quantity {i}']] = TEMP_VAL
                df.loc[df[f'AFV {i}'] == 'NULL', [f'AFV {i}']] = TEMP_VAL

                df.loc[(df[f'Quantity {i}'] > 50) & (df[f'Quantity {i}'] != TEMP_VAL), [f'Quantity {i}']] = np.nan

                for idx in df.index:

                    if isinstance(df.loc[idx, f'AFV {i}'], (int, float)):

                        if not float(df.loc[idx, f'AFV {i}']).is_integer():
                            df.loc[idx, f'AFV {i}'] = np.nan

                    else:
                        df.loc[idx, f'AFV {i}'] = np.nan


                df.loc[(df[f'Quantity {i}'] == 0) & (df[f'Food {i}'] != 'NULL'), [f'Quantity {i}']] = np.nan
                df.loc[(df[f'AFV {i}'] == 0) & (df[f'Food {i}'] != 'NULL'), [f'AFV {i}']] = np.nan

                df.loc[df[f'Quantity {i}'] == TEMP_VAL, [f'Quantity {i}']] = 'NULL'
                df.loc[df[f'AFV {i}'] == TEMP_VAL, [f'AFV {i}']] = 'NULL'

                dfCheck[f'Quantity {i}'] = df[f'Quantity {i}']
                dfCheck[f'AFV {i}'] = df[f'AFV {i}']

            except Exception:
                dfCheck[f'AFV {i}'] = [np.nan] * dfCheck.shape[0]

        return dfCheck